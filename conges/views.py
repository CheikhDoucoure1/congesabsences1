import logging

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.password_validation import validate_password
from django.contrib import messages
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.http import url_has_allowed_host_and_scheme
from django.db.models import Q, Count
from django.http import JsonResponse, FileResponse, Http404
from django.views.decorators.http import require_POST
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from .models import Employe, DemandeConge, TypeConge, SoldeConge, Notification, Departement, Recrutement, Depart, CongeSupplementaire, HistoriqueModification
from .validators import (
    valider_fichier, COULEUR_HEX_RE,
    EXTENSIONS_JUSTIFICATIF, EXTENSIONS_AVATAR, EXTENSIONS_IMPORT_EXCEL,
    TAILLE_MAX_JUSTIFICATIF, TAILLE_MAX_AVATAR, TAILLE_MAX_IMPORT_EXCEL,
)

logger = logging.getLogger(__name__)

# Attempts allowed before a (client IP, email) pair is temporarily throttled
# on the login form — mitigates online brute-force / credential stuffing.
LOGIN_MAX_TENTATIVES = 5
LOGIN_BLOCAGE_SECONDES = 300

# Roles a non-admin (rh) user is allowed to grant when creating/importing
# employees. Only an existing admin account can mint another admin or set
# someone as dg.
ROLES_ATTRIBUABLES_PAR_RH = {'employe', 'manager', 'rh'}


def _int_ou(valeur, defaut):
    """Best-effort int() that falls back instead of raising on bad input
    (e.g. a hand-edited ?annee=abc query string)."""
    try:
        return int(valeur)
    except (TypeError, ValueError):
        return defaut


def _roles_autorises_pour(user):
    if user.role == 'admin':
        return {c for c, _ in Employe.ROLE_CHOICES}
    return ROLES_ATTRIBUABLES_PAR_RH


def _demande_visible_ou_404(user, demande_id):
    """Fetch a DemandeConge, scoped to what `user` is allowed to see.

    - employe : only their own requests.
    - manager : their own requests + their direct team's (never another
      manager's team — this is the whole point of the scoping; without it
      any manager could view/act on any employee's request by guessing IDs).
    - rh / dg / admin : full visibility by design (they oversee everyone).
    """
    if user.role == 'manager':
        return get_object_or_404(
            DemandeConge,
            Q(id=demande_id) & (Q(employe__in=user.subordonnes.all()) | Q(employe=user))
        )
    if user.is_manager_or_above:  # rh, dg, admin
        return get_object_or_404(DemandeConge, id=demande_id)
    return get_object_or_404(DemandeConge, id=demande_id, employe=user)


def _cle_limitation_connexion(request, email):
    ip = request.META.get('REMOTE_ADDR', 'inconnu')
    return f'login_attempts:{ip}:{email.lower()}'


def _next_url_sure(request, valeur_brute):
    if valeur_brute and url_has_allowed_host_and_scheme(
        valeur_brute, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        return valeur_brute
    return 'tableau_de_bord'


def connexion(request):
    if request.user.is_authenticated:
        return redirect('tableau_de_bord')
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        next_url = _next_url_sure(request, request.POST.get('next') or request.GET.get('next'))

        cle = _cle_limitation_connexion(request, email)
        if cache.get(cle, 0) >= LOGIN_MAX_TENTATIVES:
            messages.error(request, "Trop de tentatives échouées. Réessayez dans quelques minutes.")
            return render(request, 'conges/connexion.html')

        # Resolve to the local `username` field ModelBackend expects when a
        # matching account already exists (demo/local accounts, previously
        # imported employees, or an AD user who has already logged in once
        # before). Otherwise pass the typed value straight through — when
        # LDAP is enabled, LDAPBackend's own search matches on email too
        # (see AUTH_LDAP_USER_SEARCH), so a brand-new AD user who has never
        # logged in here yet can still be found and auto-provisioned on
        # first success.
        try:
            identifiant = Employe.objects.get(email__iexact=email).username
        except Employe.DoesNotExist:
            identifiant = email

        user = authenticate(request, username=identifiant, password=password)
        if user and user.actif:
            cache.delete(cle)
            login(request, user)
            return redirect(next_url)
        cache.set(cle, cache.get(cle, 0) + 1, LOGIN_BLOCAGE_SECONDES)
        messages.error(request, 'Email ou mot de passe incorrect.')
    return render(request, 'conges/connexion.html')


def deconnexion(request):
    logout(request)
    return redirect('connexion')


@login_required
def tableau_de_bord(request):
    user = request.user
    annee = date.today().year

    mes_demandes_recentes = DemandeConge.objects.filter(
        employe=user
    ).select_related('type_conge')[:5]

    mes_demandes_en_attente = DemandeConge.objects.filter(
        employe=user, statut='en_attente'
    ).count()

    mes_demandes_approuvees = DemandeConge.objects.filter(
        employe=user, statut='approuve',
        date_soumission__year=annee
    ).count()

    mes_soldes = SoldeConge.objects.filter(
        employe=user, annee=annee
    ).select_related('type_conge')

    a_approuver = 0
    demandes_equipe = []
    if user.is_manager_or_above:
        if user.role == 'manager':
            equipe_ids = user.subordonnes.values_list('id', flat=True)
            a_approuver = DemandeConge.objects.filter(
                employe__in=equipe_ids, statut='en_attente'
            ).count()
        elif user.role == 'dg':
            a_approuver = DemandeConge.objects.filter(statut='validee').count()
        elif user.role in ('rh', 'admin'):
            a_approuver = DemandeConge.objects.filter(
                statut__in=['en_attente', 'validee_manager']
            ).count()

        demandes_equipe = DemandeConge.objects.filter(
            statut='approuve',
            date_fin__gte=date.today()
        ).select_related('employe', 'type_conge')[:10]

    notifications_non_lues = user.notifications.filter(lue=False).count()

    context = {
        'mes_demandes_recentes': mes_demandes_recentes,
        'mes_demandes_en_attente': mes_demandes_en_attente,
        'mes_demandes_approuvees': mes_demandes_approuvees,
        'mes_soldes': mes_soldes,
        'a_approuver': a_approuver,
        'demandes_equipe': demandes_equipe,
        'notifications_non_lues': notifications_non_lues,
        'today': date.today(),
    }
    return render(request, 'conges/tableau_de_bord.html', context)


@login_required
def nouvelle_demande(request):
    user = request.user
    types_conge = TypeConge.objects.filter(actif=True)
    annee = date.today().year
    soldes_qs = SoldeConge.objects.filter(employe=user, annee=annee)
    soldes = {s.type_conge_id: s for s in soldes_qs}
    soldes_json = {str(s.type_conge_id): float(s.jours_restants) for s in soldes_qs}
    employes_liste = Employe.objects.filter(actif=True).exclude(id=user.id).order_by('last_name', 'first_name')

    if request.method == 'POST':
        type_id = request.POST.get('type_conge')
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')
        demi_journee = request.POST.get('demi_journee') == 'on'
        periode_demi_journee = request.POST.get('periode_demi_journee', '')
        motif = request.POST.get('motif', '').strip()
        justificatif = request.FILES.get('justificatif')
        interimaire_id = request.POST.get('interimaire') or None

        errors = []
        if not type_id:
            errors.append("Veuillez sélectionner un type de congé.")
        if not date_debut_str:
            errors.append("La date de début est requise.")
        if not date_fin_str:
            errors.append("La date de fin est requise.")

        type_conge_obj = None
        if type_id:
            try:
                type_conge_obj = TypeConge.objects.get(id=type_id)
                if type_conge_obj.categorie == 'absence' and not motif:
                    errors.append("Le motif est requis pour une demande d'absence.")
            except TypeConge.DoesNotExist:
                errors.append("Type invalide.")

        if not errors:
            try:
                type_conge = type_conge_obj or TypeConge.objects.get(id=type_id)
                date_debut = date.fromisoformat(date_debut_str)
                date_fin = date.fromisoformat(date_fin_str)

                if date_debut > date_fin:
                    errors.append("La date de début doit être avant la date de fin.")
                elif date_debut < date.today():
                    errors.append("La date de début ne peut pas être dans le passé.")
                else:
                    if type_conge.necessite_justificatif and not justificatif:
                        errors.append(f"Un justificatif est requis pour '{type_conge.libelle}'.")
                    if justificatif:
                        try:
                            valider_fichier(justificatif, EXTENSIONS_JUSTIFICATIF, TAILLE_MAX_JUSTIFICATIF)
                        except ValidationError as e:
                            errors.append(e.message)
                    if not errors:
                        demande = DemandeConge(
                            employe=user,
                            type_conge=type_conge,
                            date_debut=date_debut,
                            date_fin=date_fin,
                            demi_journee=demi_journee,
                            periode_demi_journee=periode_demi_journee if demi_journee else '',
                            motif=motif,
                            interimaire_id=interimaire_id,
                        )
                        if justificatif:
                            demande.justificatif = justificatif
                        demande.save()

                        _notifier_manager(demande)

                        messages.success(request, f"Votre demande {demande.reference} a été soumise avec succès.")
                        return redirect('mes_demandes')
            except TypeConge.DoesNotExist:
                errors.append("Type de congé invalide.")
            except ValueError:
                errors.append("Format de date invalide.")

        for err in errors:
            messages.error(request, err)

    context = {
        'types_conge': types_conge,
        'soldes': soldes,
        'soldes_json': soldes_json,
        'employes_liste': employes_liste,
        'today': date.today().isoformat(),
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/nouvelle_demande.html', context)


def _notifier_manager(demande):
    employe = demande.employe
    if employe.manager:
        Notification.objects.create(
            destinataire=employe.manager,
            titre=f"Nouvelle demande de congé",
            message=f"{employe.get_full_name()} a soumis une demande de {demande.type_conge.libelle} "
                    f"du {demande.date_debut.strftime('%d/%m/%Y')} au {demande.date_fin.strftime('%d/%m/%Y')}.",
            lien=f"/approbations/{demande.id}/",
        )
    for rh in Employe.objects.filter(role__in=['rh', 'admin'], actif=True):
        Notification.objects.create(
            destinataire=rh,
            titre=f"Nouvelle demande - {employe.get_full_name()}",
            message=f"Demande {demande.reference} : {demande.type_conge.libelle} "
                    f"({demande.nombre_jours} jour(s)).",
            lien=f"/approbations/{demande.id}/",
        )


def _notifier_rh(demande):
    for rh in Employe.objects.filter(role__in=['rh', 'admin'], actif=True):
        Notification.objects.create(
            destinataire=rh,
            titre="Demande en attente de votre validation",
            message=f"Demande {demande.reference} de {demande.employe.get_full_name()} "
                    f"({demande.type_conge.libelle}, {demande.nombre_jours} jour(s)) "
                    f"a été validée par {demande.valide_par.get_full_name()} et attend votre validation.",
            lien=f"/approbations/{demande.id}/",
        )


def _notifier_dg(demande):
    validateur = demande.valide_par_rh or demande.valide_par
    for dg in Employe.objects.filter(role='dg', actif=True):
        Notification.objects.create(
            destinataire=dg,
            titre="Demande en attente de votre validation",
            message=f"Demande {demande.reference} de {demande.employe.get_full_name()} "
                    f"({demande.type_conge.libelle}, {demande.nombre_jours} jour(s)) "
                    f"a été validée par {validateur.get_full_name()} et attend votre validation finale.",
            lien=f"/approbations/{demande.id}/",
        )


@login_required
def mes_demandes(request):
    user = request.user
    statut_filtre = request.GET.get('statut', '')
    type_filtre = request.GET.get('type', '')
    annee_filtre = request.GET.get('annee', str(date.today().year))

    demandes = DemandeConge.objects.filter(employe=user).select_related('type_conge', 'traite_par')

    if statut_filtre:
        demandes = demandes.filter(statut=statut_filtre)
    if type_filtre:
        demandes = demandes.filter(type_conge__code=type_filtre)
    if annee_filtre:
        demandes = demandes.filter(date_soumission__year=annee_filtre)

    types_conge = TypeConge.objects.filter(actif=True)
    annees = range(date.today().year, date.today().year - 4, -1)

    context = {
        'demandes': demandes,
        'types_conge': types_conge,
        'statut_filtre': statut_filtre,
        'type_filtre': type_filtre,
        'annee_filtre': annee_filtre,
        'annees': annees,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/mes_demandes.html', context)


@login_required
@require_POST
def annuler_demande(request, demande_id):
    demande = get_object_or_404(DemandeConge, id=demande_id, employe=request.user)
    if demande.statut == 'en_attente':
        demande.statut = 'annule'
        demande.save()
        messages.success(request, f"La demande {demande.reference} a été annulée.")
    else:
        messages.error(request, "Cette demande ne peut plus être annulée.")
    return redirect('mes_demandes')


@login_required
def mon_solde(request):
    user = request.user
    annee = _int_ou(request.GET.get('annee'), date.today().year)
    soldes = SoldeConge.objects.filter(
        employe=user, annee=annee
    ).select_related('type_conge').order_by('type_conge__libelle')

    demandes_approuvees = DemandeConge.objects.filter(
        employe=user,
        statut='approuve',
        date_debut__year=annee
    ).select_related('type_conge').order_by('-date_debut')

    annees = range(date.today().year, date.today().year - 3, -1)

    context = {
        'soldes': soldes,
        'demandes_approuvees': demandes_approuvees,
        'annee': annee,
        'annees': annees,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/mon_solde.html', context)


@login_required
def calendrier(request):
    user = request.user
    today = date.today()
    mois = _int_ou(request.GET.get('mois'), today.month)
    annee_cal = _int_ou(request.GET.get('annee'), today.year)

    if mois < 1:
        mois = 12
        annee_cal -= 1
    elif mois > 12:
        mois = 1
        annee_cal += 1

    premier_jour = date(annee_cal, mois, 1)
    if mois == 12:
        dernier_jour = date(annee_cal + 1, 1, 1) - timedelta(days=1)
    else:
        dernier_jour = date(annee_cal, mois + 1, 1) - timedelta(days=1)

    if user.is_manager_or_above:
        demandes = DemandeConge.objects.filter(
            statut='approuve',
            date_debut__lte=dernier_jour,
            date_fin__gte=premier_jour
        ).select_related('employe', 'type_conge')
    else:
        equipe_ids = [user.id]
        if user.manager:
            equipe_ids += list(Employe.objects.filter(manager=user.manager).values_list('id', flat=True))
        demandes = DemandeConge.objects.filter(
            employe__in=equipe_ids,
            statut='approuve',
            date_debut__lte=dernier_jour,
            date_fin__gte=premier_jour
        ).select_related('employe', 'type_conge')

    evenements = []
    for d in demandes:
        evenements.append({
            'id': d.id,
            'titre': d.employe.get_full_name(),
            'type': d.type_conge.libelle,
            'couleur': d.type_conge.couleur,
            'debut': d.date_debut.isoformat(),
            'fin': d.date_fin.isoformat(),
            'jours': float(d.nombre_jours),
            'est_moi': d.employe_id == user.id,
        })

    mois_nom = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'][mois]

    context = {
        'evenements_json': evenements,
        'mois': mois,
        'annee': annee_cal,
        'mois_nom': mois_nom,
        'premier_jour': premier_jour,
        'dernier_jour': dernier_jour,
        'today': today,
        'mois_precedent': (mois - 1) or 12,
        'annee_precedent': annee_cal if mois > 1 else annee_cal - 1,
        'mois_suivant': (mois % 12) + 1,
        'annee_suivant': annee_cal if mois < 12 else annee_cal + 1,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/calendrier.html', context)


@login_required
def approbations(request):
    user = request.user
    if not user.is_manager_or_above:
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    statut_filtre = request.GET.get('statut', 'validee' if user.is_dg else 'en_attente')

    if user.role == 'manager':
        equipe_ids = user.subordonnes.values_list('id', flat=True)
        base_qs = DemandeConge.objects.filter(employe__in=equipe_ids)
    else:
        base_qs = DemandeConge.objects.all()

    demandes = base_qs.select_related('employe', 'type_conge', 'traite_par', 'valide_par', 'valide_par_rh')
    if statut_filtre:
        demandes = demandes.filter(statut=statut_filtre)

    stats = {
        'en_attente': base_qs.filter(statut='en_attente').count(),
        'validee_manager': base_qs.filter(statut='validee_manager').count(),
        'validee': base_qs.filter(statut='validee').count(),
    }

    context = {
        'demandes': demandes,
        'statut_filtre': statut_filtre,
        'stats': stats,
        'is_dg': user.is_dg,
        'is_rh': user.role in ('rh', 'admin'),
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/approbations.html', context)


@login_required
def detail_demande(request, demande_id):
    user = request.user
    demande = _demande_visible_ou_404(user, demande_id)

    context = {
        'demande': demande,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/detail_demande.html', context)


@login_required
def voir_justificatif(request, demande_id):
    """Serve a leave request's supporting document only to people who are
    allowed to see the request itself (see _demande_visible_ou_404) — the
    file is otherwise not reachable at all (it is deliberately excluded
    from the public /media/ static mount, see urls.py)."""
    demande = _demande_visible_ou_404(request.user, demande_id)
    if not demande.justificatif:
        raise Http404
    return FileResponse(
        demande.justificatif.open('rb'),
        as_attachment=True,
        filename=demande.justificatif.name.rsplit('/', 1)[-1],
    )


@login_required
def modifier_demande(request, demande_id):
    user = request.user
    if user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    demande = get_object_or_404(DemandeConge, id=demande_id)

    if demande.statut not in ('en_attente', 'validee_manager', 'validee'):
        messages.error(request, "Cette demande a déjà été traitée et ne peut plus être modifiée.")
        return redirect('detail_demande', demande_id=demande.id)

    types_conge = TypeConge.objects.filter(actif=True)
    employes_liste = Employe.objects.filter(actif=True).exclude(id=demande.employe_id).order_by('last_name', 'first_name')

    if request.method == 'POST':
        type_id = request.POST.get('type_conge')
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')
        demi_journee = request.POST.get('demi_journee') == 'on'
        periode_demi_journee = request.POST.get('periode_demi_journee', '')
        motif = request.POST.get('motif', '').strip()
        justificatif = request.FILES.get('justificatif')
        interimaire_id = request.POST.get('interimaire') or None
        nombre_jours_str = request.POST.get('nombre_jours', '').strip()

        errors = []
        type_conge_obj = None
        if not type_id:
            errors.append("Veuillez sélectionner un type de congé.")
        else:
            try:
                type_conge_obj = TypeConge.objects.get(id=type_id)
            except TypeConge.DoesNotExist:
                errors.append("Type de congé invalide.")

        date_debut = date_fin = None
        if not date_debut_str or not date_fin_str:
            errors.append("Les dates de début et de fin sont requises.")
        else:
            try:
                date_debut = date.fromisoformat(date_debut_str)
                date_fin = date.fromisoformat(date_fin_str)
                if date_debut > date_fin:
                    errors.append("La date de début doit être avant la date de fin.")
            except ValueError:
                errors.append("Format de date invalide.")

        if type_conge_obj and type_conge_obj.categorie == 'absence' and not motif:
            errors.append("Le motif est requis pour une demande d'absence.")

        nombre_jours = None
        if nombre_jours_str:
            try:
                nombre_jours = Decimal(nombre_jours_str.replace(',', '.'))
                if nombre_jours <= 0:
                    errors.append("Le nombre de jours doit être positif.")
            except InvalidOperation:
                errors.append("Nombre de jours invalide.")

        if justificatif:
            try:
                valider_fichier(justificatif, EXTENSIONS_JUSTIFICATIF, TAILLE_MAX_JUSTIFICATIF)
            except ValidationError as e:
                errors.append(e.message)

        if not errors:
            avant = {
                'Type de congé': demande.type_conge.libelle,
                'Date de début': demande.date_debut.strftime('%d/%m/%Y'),
                'Date de fin': demande.date_fin.strftime('%d/%m/%Y'),
                'Nombre de jours': str(demande.nombre_jours),
                'Motif': demande.motif or '—',
                'Intérimaire': demande.interimaire.get_full_name() if demande.interimaire else '—',
            }

            demande.type_conge = type_conge_obj
            demande.date_debut = date_debut
            demande.date_fin = date_fin
            demande.demi_journee = demi_journee
            demande.periode_demi_journee = periode_demi_journee if demi_journee else ''
            demande.motif = motif
            demande.interimaire_id = interimaire_id
            if justificatif:
                demande.justificatif = justificatif
            demande.nombre_jours = nombre_jours if nombre_jours is not None else demande.calculer_jours()
            demande.save()

            apres = {
                'Type de congé': demande.type_conge.libelle,
                'Date de début': demande.date_debut.strftime('%d/%m/%Y'),
                'Date de fin': demande.date_fin.strftime('%d/%m/%Y'),
                'Nombre de jours': str(demande.nombre_jours),
                'Motif': demande.motif or '—',
                'Intérimaire': demande.interimaire.get_full_name() if demande.interimaire else '—',
            }
            changements = [
                f"{champ} : « {avant[champ]} » → « {apres[champ]} »"
                for champ in avant if avant[champ] != apres[champ]
            ]
            HistoriqueModification.objects.create(
                type_action='demande_modifiee',
                auteur=user,
                employe_concerne=demande.employe,
                demande=demande,
                description=(
                    f"Demande {demande.reference} de {demande.employe.get_full_name()} modifiée : "
                    + ("; ".join(changements) if changements else "aucune valeur modifiée.")
                ),
            )

            messages.success(request, f"La demande {demande.reference} a été modifiée avec succès.")
            return redirect('detail_demande', demande_id=demande.id)

        for err in errors:
            messages.error(request, err)

    context = {
        'demande': demande,
        'types_conge': types_conge,
        'employes_liste': employes_liste,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/modifier_demande.html', context)


@login_required
@require_POST
def traiter_demande(request, demande_id):
    user = request.user
    if not user.is_manager_or_above:
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    if user.role == 'manager':
        # A manager may only act on their own team's requests — without this
        # filter any manager could approve/reject anyone's request by
        # guessing the demande_id in the URL.
        demande = get_object_or_404(
            DemandeConge, id=demande_id, employe__in=user.subordonnes.all()
        )
    else:
        demande = get_object_or_404(DemandeConge, id=demande_id)

    if demande.employe_id == user.id:
        messages.error(request, "Vous ne pouvez pas traiter votre propre demande.")
        return redirect('approbations')

    if demande.statut not in ('en_attente', 'validee_manager', 'validee'):
        messages.error(request, "Cette demande a déjà été traitée.")
        return redirect('approbations')

    action = request.POST.get('action')
    commentaire = request.POST.get('commentaire', '').strip()

    if action not in ('approuver', 'rejeter'):
        messages.error(request, "Action invalide.")
        return redirect('approbations')

    if not user.is_dg:
        if user.role == 'manager' and demande.statut != 'en_attente':
            messages.error(request, "Cette demande n'est plus en attente de votre validation.")
            return redirect('approbations')
        if user.role in ('rh', 'admin') and demande.statut not in ('en_attente', 'validee_manager'):
            messages.error(request, "Cette demande est en attente de la décision du DG.")
            return redirect('approbations')

    if action == 'rejeter':
        demande.statut = 'rejete'
        demande.date_traitement = timezone.now()
        demande.traite_par = user
        demande.commentaire_traitement = commentaire
        demande.save()
        _notifier_employe(demande, 'rejete', commentaire)
        messages.success(request, f"La demande {demande.reference} a été rejetée.")
        return redirect('approbations')

    if user.is_dg:
        demande.statut = 'approuve'
        demande.date_traitement = timezone.now()
        demande.traite_par = user
        demande.commentaire_traitement = commentaire
        demande.save()
        _mettre_a_jour_solde(demande)
        _notifier_employe(demande, 'approuve', commentaire)
        messages.success(request, f"La demande {demande.reference} a été définitivement approuvée.")
    elif user.role == 'manager':
        demande.statut = 'validee_manager'
        demande.valide_par = user
        demande.date_validation = timezone.now()
        demande.commentaire_validation = commentaire
        demande.save()
        _notifier_rh(demande)
        _notifier_employe(demande, 'validee_manager', commentaire)
        messages.success(request, f"La demande {demande.reference} a été validée et transmise au RH.")
    else:
        demande.statut = 'validee'
        demande.valide_par_rh = user
        demande.date_validation_rh = timezone.now()
        demande.commentaire_validation_rh = commentaire
        demande.save()
        _notifier_dg(demande)
        _notifier_employe(demande, 'validee', commentaire)
        messages.success(request, f"La demande {demande.reference} a été validée et transmise au DG pour approbation finale.")

    return redirect('approbations')


def _mettre_a_jour_solde(demande):
    annee = demande.date_debut.year
    solde, _ = SoldeConge.objects.get_or_create(
        employe=demande.employe,
        type_conge=demande.type_conge,
        annee=annee,
        defaults={'jours_acquis': demande.type_conge.jours_max}
    )
    solde.jours_pris += demande.nombre_jours
    solde.save()


def _notifier_employe(demande, statut, commentaire):
    labels = {
        'validee_manager': "validée par votre manager, en attente de la validation du RH",
        'validee': "validée par le RH, en attente de la validation finale du DG",
        'approuve': "définitivement approuvée",
        'rejete': "rejetée",
    }
    statut_label = labels.get(statut, statut)
    msg = f"Votre demande {demande.reference} ({demande.type_conge.libelle}) a été {statut_label}."
    if commentaire:
        msg += f" Commentaire : {commentaire}"
    Notification.objects.create(
        destinataire=demande.employe,
        titre=f"Demande {statut_label}",
        message=msg,
        lien=f"/mes-demandes/",
    )


@login_required
def equipe(request):
    user = request.user
    if not user.is_manager_or_above:
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    if user.role == 'manager':
        employes = Employe.objects.filter(manager=user, actif=True).select_related('departement')
    else:
        employes = Employe.objects.filter(actif=True).select_related('departement', 'manager')

    annee = date.today().year
    aujourd_hui = date.today()
    absents_aujourd_hui = DemandeConge.objects.filter(
        statut='approuve',
        date_debut__lte=aujourd_hui,
        date_fin__gte=aujourd_hui,
    ).select_related('employe', 'type_conge')

    if user.role == 'manager':
        equipe_ids = user.subordonnes.values_list('id', flat=True)
        absents_aujourd_hui = absents_aujourd_hui.filter(employe__in=equipe_ids)

    context = {
        'employes': employes,
        'absents_aujourd_hui': absents_aujourd_hui,
        'aujourd_hui': aujourd_hui,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/equipe.html', context)


@login_required
def historique_modifications(request):
    user = request.user
    if not user.is_manager_or_above:
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    historique = HistoriqueModification.objects.select_related(
        'auteur', 'employe_concerne', 'demande'
    )
    if user.role == 'manager':
        equipe_ids = list(user.subordonnes.values_list('id', flat=True))
        historique = historique.filter(employe_concerne_id__in=equipe_ids)

    context = {
        'historique': historique[:300],
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/historique_modifications.html', context)


@login_required
def administration(request):
    user = request.user
    if user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    onglet = request.GET.get('onglet', 'employes')
    employes = Employe.objects.filter(actif=True).select_related('departement').order_by('last_name')
    departements = Departement.objects.all()
    types_conge = TypeConge.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_employe':
            return _ajouter_employe(request)
        elif action == 'add_type_conge':
            return _ajouter_type_conge(request)
        elif action == 'init_soldes':
            return _initialiser_soldes(request)
        elif action == 'add_conge_supplementaire':
            return _ajouter_conge_supplementaire(request)
        elif action == 'modifier_solde':
            return _modifier_solde(request)

    annee_soldes = _int_ou(request.GET.get('annee'), date.today().year)
    soldes_tous = SoldeConge.objects.filter(annee=annee_soldes).select_related(
        'employe', 'type_conge'
    ).order_by('employe__last_name', 'employe__first_name', 'type_conge__libelle')

    context = {
        'onglet': onglet,
        'employes': employes,
        'departements': departements,
        'types_conge': types_conge,
        'soldes_tous': soldes_tous,
        'annee_soldes': annee_soldes,
        'conges_supplementaires': CongeSupplementaire.objects.select_related(
            'employe', 'type_conge', 'accorde_par'
        )[:50],
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/administration.html', context)


def _creer_employe_depuis_post(request):
    dept_id = request.POST.get('departement')
    manager_id = request.POST.get('manager')
    username = request.POST.get('email', '').split('@')[0]

    role_demande = request.POST.get('role', 'employe')
    if role_demande not in _roles_autorises_pour(request.user):
        # e.g. an 'rh' account trying to grant 'admin'/'dg' — silently fall
        # back instead of trusting client-submitted privilege escalation.
        role_demande = 'employe'

    mot_de_passe_fourni = request.POST.get('password')
    mot_de_passe = mot_de_passe_fourni or get_random_string(12)

    emp = Employe.objects.create_user(
        username=username,
        email=request.POST.get('email'),
        password=mot_de_passe,
        first_name=request.POST.get('prenom', ''),
        last_name=request.POST.get('nom', ''),
    )
    emp.poste = request.POST.get('poste', '')
    emp.role = role_demande
    emp.matricule = request.POST.get('matricule', '')
    if dept_id:
        emp.departement_id = dept_id
    if manager_id:
        emp.manager_id = manager_id
    emp.save()
    _creer_soldes_employe(emp)
    return emp, mot_de_passe, not mot_de_passe_fourni


def _ajouter_employe(request):
    try:
        emp, mot_de_passe, generee = _creer_employe_depuis_post(request)
        msg = f"L'employé {emp.get_full_name()} a été créé avec succès."
        if generee:
            msg += f" Mot de passe généré : {mot_de_passe} (à transmettre à l'employé de façon sécurisée)."
        messages.success(request, msg)
    except Exception:
        logger.exception("Échec de création d'employé via l'administration")
        messages.error(request, "Erreur lors de la création de l'employé. Vérifiez les informations saisies.")
    return redirect('/administration/?onglet=employes')


def _creer_soldes_employe(employe):
    annee = date.today().year
    for type_conge in TypeConge.objects.filter(actif=True):
        SoldeConge.objects.get_or_create(
            employe=employe,
            type_conge=type_conge,
            annee=annee,
            defaults={'jours_acquis': type_conge.jours_max}
        )


def _ajouter_type_conge(request):
    try:
        couleur = request.POST.get('couleur', '#2196F3')
        if not COULEUR_HEX_RE.match(couleur):
            raise ValueError("Couleur invalide (format attendu : #RRGGBB).")
        TypeConge.objects.create(
            code=request.POST.get('code', ''),
            libelle=request.POST.get('libelle', ''),
            couleur=couleur,
            jours_max=_int_ou(request.POST.get('jours_max'), 30),
            necessite_justificatif=request.POST.get('necessite_justificatif') == 'on',
            description=request.POST.get('description', ''),
        )
        messages.success(request, "Type de congé ajouté avec succès.")
    except ValueError as e:
        messages.error(request, str(e))
    except Exception:
        logger.exception("Échec de création de type de congé")
        messages.error(request, "Erreur lors de la création du type de congé.")
    return redirect('/administration/?onglet=types_conge')


def _initialiser_soldes(request):
    annee = _int_ou(request.POST.get('annee'), date.today().year)
    count = 0
    for emp in Employe.objects.filter(actif=True):
        for tc in TypeConge.objects.filter(actif=True):
            _, created = SoldeConge.objects.get_or_create(
                employe=emp, type_conge=tc, annee=annee,
                defaults={'jours_acquis': tc.jours_max}
            )
            if created:
                count += 1
    messages.success(request, f"{count} soldes initialisés pour {annee}.")
    return redirect('/administration/?onglet=soldes')


def _modifier_solde(request):
    annee_filtre = request.POST.get('annee_filtre') or date.today().year
    try:
        solde = SoldeConge.objects.select_related('employe', 'type_conge').get(
            id=request.POST.get('solde_id')
        )

        def _lire_decimal(champ):
            try:
                return Decimal(request.POST.get(champ, '0').replace(',', '.'))
            except InvalidOperation:
                raise ValueError(f"Valeur invalide pour « {champ} ».")

        nouvelles_valeurs = {
            'jours_acquis': _lire_decimal('jours_acquis'),
            'jours_pris': _lire_decimal('jours_pris'),
            'jours_reportes': _lire_decimal('jours_reportes'),
            'jours_supplementaires': _lire_decimal('jours_supplementaires'),
        }
        for champ, valeur in nouvelles_valeurs.items():
            if valeur < 0:
                raise ValueError(f"« {champ} » ne peut pas être négatif.")

        labels = {
            'jours_acquis': 'Jours acquis',
            'jours_pris': 'Jours pris',
            'jours_reportes': 'Jours reportés',
            'jours_supplementaires': 'Jours supplémentaires',
        }
        avant = {champ: getattr(solde, champ) for champ in nouvelles_valeurs}

        for champ, valeur in nouvelles_valeurs.items():
            setattr(solde, champ, valeur)
        solde.save()

        changements = [
            f"{labels[champ]} : « {avant[champ]} » → « {nouvelles_valeurs[champ]} »"
            for champ in nouvelles_valeurs if avant[champ] != nouvelles_valeurs[champ]
        ]
        if changements:
            HistoriqueModification.objects.create(
                type_action='solde_modifie',
                auteur=request.user,
                employe_concerne=solde.employe,
                description=(
                    f"Solde {solde.type_conge.libelle} ({solde.annee}) de {solde.employe.get_full_name()} modifié : "
                    + "; ".join(changements)
                ),
            )
        messages.success(
            request,
            f"Solde de {solde.employe.get_full_name()} ({solde.type_conge.libelle}, {solde.annee}) mis à jour."
        )
    except SoldeConge.DoesNotExist:
        messages.error(request, "Solde introuvable.")
    except ValueError as e:
        messages.error(request, str(e))
    except Exception:
        logger.exception("Échec de modification de solde")
        messages.error(request, "Erreur lors de la mise à jour du solde.")
    return redirect(f"/administration/?onglet=soldes&annee={annee_filtre}")


def _ajouter_conge_supplementaire(request):
    try:
        employe = Employe.objects.get(id=request.POST.get('employe'))
        type_conge = TypeConge.objects.get(id=request.POST.get('type_conge'))
        annee = _int_ou(request.POST.get('annee'), date.today().year)
        try:
            jours = Decimal(request.POST.get('nombre_jours', '').replace(',', '.'))
        except InvalidOperation:
            raise ValueError("Nombre de jours invalide.")
        if jours <= 0:
            raise ValueError("Le nombre de jours doit être positif.")
        motif = request.POST.get('motif', '').strip()

        CongeSupplementaire.objects.create(
            employe=employe, type_conge=type_conge, annee=annee,
            nombre_jours=jours, motif=motif, accorde_par=request.user,
        )
        solde, _ = SoldeConge.objects.get_or_create(
            employe=employe, type_conge=type_conge, annee=annee,
            defaults={'jours_acquis': type_conge.jours_max}
        )
        solde.jours_supplementaires += jours
        solde.save()
        HistoriqueModification.objects.create(
            type_action='conge_supplementaire',
            auteur=request.user,
            employe_concerne=employe,
            description=(
                f"{jours} jour(s) de congé supplémentaire ({type_conge.libelle}, {annee}) "
                f"accordé(s) à {employe.get_full_name()}."
                + (f" Motif : {motif}" if motif else "")
            ),
        )
        Notification.objects.create(
            destinataire=employe,
            titre="Congé supplémentaire crédité",
            message=(
                f"{jours} jour(s) de congé supplémentaire ({type_conge.libelle}, {annee}) "
                f"ont été crédités sur votre solde."
                + (f" Motif : {motif}" if motif else "")
            ),
            lien="/mon-solde/",
        )
        messages.success(
            request,
            f"{jours} jour(s) supplémentaire(s) accordé(s) à {employe.get_full_name()} "
            f"({type_conge.libelle}, {annee})."
        )
    except ValueError as e:
        messages.error(request, str(e))
    except (Employe.DoesNotExist, TypeConge.DoesNotExist):
        messages.error(request, "Employé ou type de congé introuvable.")
    except Exception:
        logger.exception("Échec d'attribution de congé supplémentaire")
        messages.error(request, "Erreur lors de l'attribution du congé supplémentaire.")
    return redirect('/administration/?onglet=conges_supplementaires')


@login_required
def importer_employes(request):
    if request.user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('administration')

    if request.method != 'POST' or not request.FILES.get('fichier_excel'):
        messages.error(request, "Aucun fichier fourni.")
        return redirect('/administration/?onglet=employes')

    import openpyxl
    fichier = request.FILES['fichier_excel']
    try:
        valider_fichier(fichier, EXTENSIONS_IMPORT_EXCEL, TAILLE_MAX_IMPORT_EXCEL)
    except ValidationError as e:
        messages.error(request, e.message)
        return redirect('/administration/?onglet=employes')

    try:
        # read_only keeps memory use bounded for large sheets; data_only
        # reads computed values instead of formulas.
        wb = openpyxl.load_workbook(fichier, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        messages.error(request, "Fichier Excel invalide.")
        return redirect('/administration/?onglet=employes')

    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required = {'prenom', 'nom', 'email'}
    if not required.issubset(set(headers)):
        messages.error(request, "Colonnes obligatoires manquantes : Prenom, Nom, Email.")
        return redirect('/administration/?onglet=employes')

    def col(row, name):
        try:
            idx = headers.index(name)
            val = row[idx].value
            return str(val).strip() if val is not None else ''
        except (ValueError, IndexError):
            return ''

    roles_autorises = _roles_autorises_pour(request.user)
    crees, mis_a_jour, erreurs, comptes_generes = 0, 0, [], []
    annee = date.today().year
    type_annuel = TypeConge.objects.filter(code='annuel').first()

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if not any(c.value for c in row):
            continue
        prenom = col(row, 'prenom')
        nom    = col(row, 'nom')
        email  = col(row, 'email')
        if not prenom or not nom or not email:
            erreurs.append(f"Ligne {i} : prénom, nom ou email manquant.")
            continue

        matricule   = col(row, 'matricule')
        poste       = col(row, 'poste')
        role        = col(row, 'role') or 'employe'
        dept_code   = col(row, 'departement')
        manager_email = col(row, 'manager')
        password_col  = col(row, 'mot de passe')

        if role not in ('employe', 'manager', 'rh', 'admin', 'dg'):
            role = 'employe'
        if role not in roles_autorises:
            # rh accounts may not grant admin/dg via a spreadsheet either.
            role = 'employe'

        password = password_col or get_random_string(12)

        dept = None
        if dept_code:
            dept = Departement.objects.filter(
                Q(code__iexact=dept_code) | Q(nom__iexact=dept_code)
            ).first()

        manager = None
        if manager_email:
            manager = Employe.objects.filter(email=manager_email).first()

        try:
            if Employe.objects.filter(email=email).exists():
                emp = Employe.objects.get(email=email)
                emp.first_name = prenom
                emp.last_name  = nom
                emp.poste      = poste
                emp.role       = role
                if dept:
                    emp.departement = dept
                if manager:
                    emp.manager = manager
                if matricule:
                    emp.matricule = matricule
                emp.actif = True
                emp.save()
                mis_a_jour += 1
            else:
                username = email.split('@')[0]
                if Employe.objects.filter(username=username).exists():
                    username = email.replace('@', '_').replace('.', '_')
                emp = Employe.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=prenom,
                    last_name=nom,
                )
                emp.poste       = poste
                emp.role        = role
                emp.departement = dept
                emp.manager     = manager
                emp.matricule   = matricule
                emp.actif       = True
                emp.save()
                if type_annuel:
                    SoldeConge.objects.get_or_create(
                        employe=emp, type_conge=type_annuel, annee=annee,
                        defaults={'jours_acquis': 24}
                    )
                crees += 1
                if not password_col:
                    comptes_generes.append((email, password))
        except Exception:
            logger.exception("Échec d'import de la ligne %s (%s)", i, email)
            erreurs.append(f"Ligne {i} ({email}) : données invalides ou déjà utilisées.")

    msg = f"Import terminé : {crees} créé(s), {mis_a_jour} mis à jour."
    if comptes_generes:
        apercu = "; ".join(f"{e} / {p}" for e, p in comptes_generes[:5])
        suite = " (…)" if len(comptes_generes) > 5 else ""
        msg += (
            f" Mot de passe généré pour {len(comptes_generes)} nouveau(x) compte(s) sans mot de passe "
            f"fourni dans le fichier : {apercu}{suite}. Transmettez-les individuellement et demandez "
            f"leur changement dès la première connexion."
        )
    if erreurs:
        msg += f" {len(erreurs)} erreur(s) : " + " | ".join(erreurs[:5])
        messages.warning(request, msg)
    else:
        messages.success(request, msg)

    return redirect('/administration/?onglet=employes')


@login_required
def telecharger_template_employes(request):
    if request.user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('administration')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employés"

    headers = ['Prenom', 'Nom', 'Email', 'Matricule', 'Poste',
               'Departement', 'Role', 'Manager', 'Mot de passe']
    widths  = [15, 15, 30, 12, 30, 30, 12, 30, 18]

    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = w

    exemples = [
        ['Aminata', 'Diallo', 'a.diallo@petrosen.sn', 'PET-001', 'Ingénieure Process',
         'DEP', 'employe', 'i.diop@petrosen.sn', ''],
        ['Moussa', 'Ndiaye', 'm.ndiaye@petrosen.sn', 'PET-002', 'Comptable',
         'FIN', 'employe', 'o.kane@petrosen.sn', ''],
    ]
    for r_idx, row in enumerate(exemples, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Feuille aide
    ws2 = wb.create_sheet("Aide")
    ws2['A1'] = "Valeurs acceptées pour la colonne Role :"
    ws2['A2'] = "employe  |  manager  |  rh  |  admin"
    ws2['A3'] = "(un compte RH important le fichier ne peut pas s'attribuer ou attribuer 'admin'/'dg' — réservé aux administrateurs)"
    ws2['A5'] = "Colonne Departement : code (DG, DRH, DEP, FIN, HSE, JUR, DSI, LOG) ou nom complet"
    ws2['A7'] = "Colonne Manager : email du manager direct (doit exister dans le système)"
    ws2['A9'] = "Mot de passe : laissez vide pour générer un mot de passe aléatoire sécurisé par employé (affiché après import, à transmettre individuellement)"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_employes_petrosen.xlsx"'
    wb.save(response)
    return response


@login_required
def notifications(request):
    user = request.user
    notifs = user.notifications.all()[:50]
    user.notifications.filter(lue=False).update(lue=True)
    context = {
        'notifs': notifs,
        'notifications_non_lues': 0,
    }
    return render(request, 'conges/notifications.html', context)


@login_required
def api_notifications(request):
    notifs = request.user.notifications.filter(lue=False)[:10]
    data = [{
        'id': n.id,
        'titre': n.titre,
        'message': n.message,
        'lien': n.lien,
        'date': n.date_creation.strftime('%d/%m/%Y %H:%M'),
    } for n in notifs]
    return JsonResponse({'notifications': data, 'count': len(data)})


@login_required
def profil(request):
    user = request.user
    if request.method == 'POST':
        user.first_name = request.POST.get('prenom', user.first_name)
        user.last_name = request.POST.get('nom', user.last_name)
        user.telephone = request.POST.get('telephone', user.telephone)

        avatar_upload = request.FILES.get('avatar')
        if avatar_upload:
            try:
                valider_fichier(avatar_upload, EXTENSIONS_AVATAR, TAILLE_MAX_AVATAR)
                user.avatar = avatar_upload
            except ValidationError as e:
                messages.error(request, e.message)

        new_pwd = request.POST.get('nouveau_mot_de_passe')
        if new_pwd:
            current_pwd = request.POST.get('mot_de_passe_actuel')
            if user.check_password(current_pwd):
                try:
                    validate_password(new_pwd, user=user)
                except ValidationError as e:
                    for err in e.messages:
                        messages.error(request, err)
                else:
                    user.set_password(new_pwd)
                    update_session_auth_hash(request, user)
                    messages.success(request, "Mot de passe modifié.")
            else:
                messages.error(request, "Mot de passe actuel incorrect.")
        user.save()
        messages.success(request, "Profil mis à jour.")
        return redirect('profil')

    context = {
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/profil.html', context)


@login_required
def recrutements(request):
    user = request.user
    if user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    if request.method == 'POST' and request.POST.get('action') == 'add_recrutement':
        return _ajouter_recrutement(request)

    filtre_statut = request.GET.get('statut', '')
    embauches = Recrutement.objects.select_related('employe', 'employe__departement', 'responsable_rh')
    if filtre_statut:
        embauches = embauches.filter(statut=filtre_statut)

    context = {
        'embauches': embauches,
        'employes': Employe.objects.order_by('last_name', 'first_name'),
        'departements': Departement.objects.all(),
        'filtre_statut': filtre_statut,
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/recrutements.html', context)


def _ajouter_recrutement(request):
    try:
        date_embauche = date.fromisoformat(request.POST.get('date_embauche'))
        mot_de_passe_genere = None
        if request.POST.get('mode_employe') == 'nouveau':
            employe, mot_de_passe, generee = _creer_employe_depuis_post(request)
            employe.date_embauche = date_embauche
            employe.save()
            if generee:
                mot_de_passe_genere = mot_de_passe
        else:
            employe = Employe.objects.get(id=request.POST.get('employe'))

        periode_essai_fin = request.POST.get('periode_essai_fin')
        Recrutement.objects.create(
            employe=employe,
            type_contrat=request.POST.get('type_contrat', 'cdi'),
            source=request.POST.get('source', ''),
            date_embauche=date_embauche,
            periode_essai_fin=date.fromisoformat(periode_essai_fin) if periode_essai_fin else None,
            responsable_rh=request.user,
        )
        msg = f"Recrutement de {employe.get_full_name()} enregistré avec succès."
        if mot_de_passe_genere:
            msg += f" Mot de passe généré : {mot_de_passe_genere} (à transmettre à l'employé de façon sécurisée)."
        messages.success(request, msg)
    except (Employe.DoesNotExist, TypeError, ValueError):
        messages.error(request, "Données invalides pour la création du recrutement.")
    except Exception:
        logger.exception("Échec de création de recrutement")
        messages.error(request, "Erreur lors de la création du recrutement.")
    return redirect('recrutements')


@login_required
@require_POST
def maj_recrutement(request, recrutement_id):
    user = request.user
    if user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    embauche = get_object_or_404(Recrutement, id=recrutement_id)
    embauche.statut = request.POST.get('statut', embauche.statut)
    embauche.contrat_signe = request.POST.get('contrat_signe') == 'on'
    embauche.visite_medicale_effectuee = request.POST.get('visite_medicale_effectuee') == 'on'
    embauche.dossier_complet = request.POST.get('dossier_complet') == 'on'
    embauche.notes = request.POST.get('notes', embauche.notes)
    embauche.save()
    messages.success(request, f"Le recrutement de {embauche.employe.get_full_name()} a été mis à jour.")
    return redirect('recrutements')


@login_required
def departs(request):
    user = request.user
    if user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    if request.method == 'POST' and request.POST.get('action') == 'add_depart':
        return _ajouter_depart(request)

    liste_departs = Depart.objects.select_related('employe', 'employe__departement', 'traite_par')

    context = {
        'liste_departs': liste_departs,
        'employes_actifs': Employe.objects.filter(actif=True).order_by('last_name', 'first_name'),
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    return render(request, 'conges/departs.html', context)


def _ajouter_depart(request):
    try:
        Depart.objects.create(
            employe_id=request.POST.get('employe'),
            type_depart=request.POST.get('type_depart', 'demission'),
            date_depart=date.fromisoformat(request.POST.get('date_depart')),
            preavis_jours=request.POST.get('preavis_jours') or None,
            motif=request.POST.get('motif', ''),
            traite_par=request.user,
        )
        messages.success(request, "Départ enregistré avec succès.")
    except (TypeError, ValueError):
        messages.error(request, "Données invalides (vérifiez la date de départ).")
    except Exception:
        logger.exception("Échec d'enregistrement de départ")
        messages.error(request, "Erreur lors de l'enregistrement du départ.")
    return redirect('departs')


@login_required
@require_POST
def maj_depart(request, depart_id):
    user = request.user
    if user.role not in ('rh', 'admin'):
        messages.error(request, "Accès non autorisé.")
        return redirect('tableau_de_bord')

    depart = get_object_or_404(Depart, id=depart_id)
    depart.statut = request.POST.get('statut', depart.statut)
    depart.entretien_sortie_effectue = request.POST.get('entretien_sortie_effectue') == 'on'
    depart.solde_tout_compte_effectue = request.POST.get('solde_tout_compte_effectue') == 'on'
    depart.materiel_restitue = request.POST.get('materiel_restitue') == 'on'
    depart.commentaire = request.POST.get('commentaire', depart.commentaire)
    depart.traite_par = request.user
    depart.save()
    messages.success(request, f"Le départ de {depart.employe.get_full_name()} a été mis à jour.")
    return redirect('departs')
